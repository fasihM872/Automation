"""Run email and WhatsApp promotion sends from configured lead files."""
import argparse
import copy
import csv
import os
import secrets
import time
from contextlib import nullcontext
from dataclasses import dataclass
from datetime import datetime
from itertools import cycle
from pathlib import Path

from dotenv import load_dotenv

import config
from content import build_message
from senders import EmailSender, WhatsAppSender, normalize_phone


@dataclass
class Business:
    name: str
    email: str = ""
    phone: str = ""
    address: str = ""


SENT_STATUSES = {"sent", "skipped", "disabled", "no_recipient"}
FAILED_STATUSES = {"failed", "not_configured"}
TRACKING_LOG = config.DATA_DIR / "email_tracking.csv"


def _pick(row, logical_name):
    headers = config.COLUMN_MAP[logical_name]
    for key, value in row.items():
        normalized_key = (key or "").strip().lower()
        if any(header in normalized_key for header in headers):
            return "" if value is None else str(value).strip()
    return ""


def load_leads(path):
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Lead file not found: {path}")

    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        yield from load_xlsx_leads(path)
        return

    if path.suffix.lower() != ".csv":
        raise ValueError(f"Unsupported lead file type: {path.suffix}. Use .csv or .xlsx")

    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            yield from _business_from_row(row)


def load_xlsx_leads(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Install openpyxl to read Excel files: pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = ["" if value is None else str(value).strip() for value in next(rows, [])]
    for values in rows:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        yield from _business_from_row(row)


def _business_from_row(row):
    business = Business(
        name=_pick(row, "name") or "there",
        email=_pick(row, "email"),
        phone=_pick(row, "phone"),
        address=_pick(row, "address"),
    )
    if business.email or business.phone:
        yield business


def load_sent(path):
    if not path.exists():
        return set()
    with open(path, newline="", encoding="utf-8") as fh:
        return {
            (row.get("niche", ""), row.get("email", ""), row.get("phone", ""))
            for row in csv.DictReader(fh)
            if row.get("email_status") in SENT_STATUSES
            and row.get("whatsapp_status") in SENT_STATUSES
        }


def append_sent(path, niche, business, message, email_status, whatsapp_status):
    path.parent.mkdir(parents=True, exist_ok=True)
    exists = path.exists()
    with open(path, "a", newline="", encoding="utf-8") as fh:
        fieldnames = [
            "sent_at",
            "niche",
            "name",
            "email",
            "phone",
            "template_name",
            "template_url",
            "email_status",
            "whatsapp_status",
        ]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerow(
            {
                "sent_at": datetime.now().isoformat(timespec="seconds"),
                "niche": niche,
                "name": business.name,
                "email": business.email,
                "phone": business.phone,
                "template_name": message.template_name,
                "template_url": message.template_url,
                "email_status": email_status,
                "whatsapp_status": whatsapp_status,
            }
        )


def _tracking_base_url(args):
    base_url = args.tracking_base_url or os.getenv("TRACKING_BASE_URL", "")
    return base_url.strip().rstrip("/")


def _tracking_url(base_url, tracking_id):
    if not base_url:
        return ""
    return f"{base_url}/track/open/{tracking_id}.gif"


def append_tracking(path, tracking_id, niche, business, message):
    path.parent.mkdir(parents=True, exist_ok=True)
    exists = path.exists()
    with open(path, "a", newline="", encoding="utf-8") as fh:
        fieldnames = [
            "tracking_id",
            "sent_at",
            "niche",
            "name",
            "email",
            "phone",
            "template_name",
            "template_url",
            "subject",
        ]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        if not exists:
            writer.writeheader()
        writer.writerow(
            {
                "tracking_id": tracking_id,
                "sent_at": datetime.now().isoformat(timespec="seconds"),
                "niche": niche,
                "name": business.name,
                "email": business.email,
                "phone": business.phone,
                "template_name": message.template_name,
                "template_url": message.template_url,
                "subject": message.subject,
            }
        )


def make_email_sender():
    return EmailSender(
        host=os.getenv("SMTP_HOST"),
        port=os.getenv("SMTP_PORT", "587"),
        username=os.getenv("SMTP_USERNAME"),
        password=os.getenv("SMTP_PASSWORD"),
        sender_name=os.getenv("SENDER_NAME", "FRZ Energy"),
        sender_email=os.getenv("SENDER_EMAIL", os.getenv("SMTP_USERNAME", "")),
        reply_to=os.getenv("REPLY_TO"),
    )


def make_whatsapp_sender():
    sid = os.getenv("TWILIO_ACCOUNT_SID")
    token = os.getenv("TWILIO_AUTH_TOKEN")
    from_number = os.getenv("TWILIO_WHATSAPP_FROM")
    if not (sid and token and from_number):
        return None
    return WhatsAppSender(sid, token, from_number, config.DEFAULT_COUNTRY_CODE)


def validate_setup(niche_name, niche_cfg, sheet):
    errors = []
    if not Path(sheet).exists():
        errors.append(f"Lead file does not exist: {sheet}")
    if not niche_cfg.get("templates"):
        errors.append(f"Niche has no templates: {niche_name}")
    for index, template in enumerate(niche_cfg.get("templates", []), start=1):
        if not template.get("name"):
            errors.append(f"Template #{index} is missing name")
        if not template.get("url"):
            errors.append(f"Template #{index} is missing url")
        preview = template.get("preview_image", "")
        if preview and not preview.startswith(("http://", "https://")):
            preview_path = Path(preview)
            if not preview_path.is_absolute():
                preview_path = config.BASE_DIR / preview_path
            if not preview_path.exists():
                errors.append(f"Preview image not found: {preview_path}")

    if not args_env_present_for_email() and not os.getenv("ALLOW_MISSING_SMTP"):
        errors.append("SMTP is not configured yet. Fill .env before real email sends.")

    if not args_env_present_for_whatsapp() and not os.getenv("ALLOW_MISSING_TWILIO"):
        errors.append("Twilio is not configured yet. Fill .env before real WhatsApp sends.")

    return errors


def args_env_present_for_email():
    return all(os.getenv(name) for name in ("SMTP_HOST", "SMTP_USERNAME", "SMTP_PASSWORD"))


def args_env_present_for_whatsapp():
    return all(os.getenv(name) for name in ("TWILIO_ACCOUNT_SID", "TWILIO_AUTH_TOKEN", "TWILIO_WHATSAPP_FROM"))


def write_preview(message, business, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = "".join(ch if ch.isalnum() else "-" for ch in business.name.lower()).strip("-") or "preview"
    path = out_dir / f"{filename}.html"
    path.write_text(message.html_body, encoding="utf-8")
    return path


def apply_message_overrides(niche_cfg, args):
    niche_cfg = copy.deepcopy(niche_cfg)
    if args.email_subject:
        niche_cfg["email_subject"] = args.email_subject
    if args.email_intro:
        niche_cfg["email_intro"] = args.email_intro
    if args.template_url:
        for template in niche_cfg.get("templates", []):
            template["url"] = args.template_url
    if args.preview_image is not None:
        for template in niche_cfg.get("templates", []):
            template["preview_image"] = args.preview_image
    return niche_cfg


def run(args):
    load_dotenv(override=True)
    niche_name = args.niche or config.ACTIVE_NICHE
    niche_cfg = apply_message_overrides(config.NICHES[niche_name], args)
    sheet = Path(args.sheet) if args.sheet else niche_cfg["sheet"]
    limit = args.limit if args.limit is not None else config.MAX_PER_RUN
    dry_run = args.dry_run or not args.send

    setup_errors = validate_setup(niche_name, niche_cfg, sheet)
    if args.check:
        print(f"Niche: {niche_name}")
        print(f"Lead file: {sheet}")
        print(f"Mode: {'send' if args.send else 'dry-run'}")
        if setup_errors:
            print("\nSetup warnings:")
            for error in setup_errors:
                print(f"- {error}")
        else:
            print("Setup looks ready.")
        if not args.send:
            return
    elif args.send:
        blocking = [
            error
            for error in setup_errors
            if "SMTP is not configured" in error and not args.no_email
            or "Twilio is not configured" in error and not args.no_whatsapp
            or "Lead file" in error
            or "Template" in error
            or "Preview image" in error
        ]
        if blocking:
            raise RuntimeError("Cannot send yet:\n" + "\n".join(f"- {error}" for error in blocking))

    leads = list(load_leads(sheet))
    if limit is not None:
        leads = leads[:limit]
    if not leads:
        print("No leads found.")
        return

    sent = load_sent(config.SENT_LOG)
    templates = cycle(niche_cfg["templates"])
    sender_name = os.getenv("SENDER_NAME", "FRZ Energy")
    sender_email = os.getenv("SENDER_EMAIL", os.getenv("SMTP_USERNAME", "info@frzenergy.store"))
    tracking_base_url = _tracking_base_url(args)

    email_sender = None if args.no_email or dry_run else make_email_sender()
    whatsapp_sender = None if args.no_whatsapp or dry_run else make_whatsapp_sender()

    with email_sender if email_sender else nullcontext() as email_client:
        for index, business in enumerate(leads, start=1):
            key = (niche_name, business.email, business.phone)
            if key in sent and not args.resend:
                print(f"SKIP already sent: {business.name} {business.email} {business.phone}")
                continue

            template = next(templates)
            tracking_id = secrets.token_urlsafe(18)
            tracking_pixel_url = _tracking_url(tracking_base_url, tracking_id)
            message = build_message(
                business,
                template,
                niche_cfg,
                niche_name,
                sender_name,
                sender_email,
                tracking_pixel_url=tracking_pixel_url,
            )

            email_status = "skipped"
            whatsapp_status = "skipped"

            if args.preview:
                preview_path = write_preview(message, business, config.DATA_DIR / "previews")
                print(f"Preview written: {preview_path}")

            if dry_run:
                print(f"\nDRY RUN #{index}: {business.name}")
                print(f"Email: {business.email}")
                print(f"Phone: {normalize_phone(business.phone, config.DEFAULT_COUNTRY_CODE) or business.phone}")
                print(f"Subject: {message.subject}")
                print(f"WhatsApp:\n{message.whatsapp_text}")
                continue

            if args.no_email:
                email_status = "disabled"
            elif not business.email:
                email_status = "no_recipient"
            elif email_client:
                try:
                    email_client.send(
                        business.email,
                        message.subject,
                        message.html_body,
                        message.text_body,
                        message.inline_images,
                    )
                    email_status = "sent"
                except Exception as exc:
                    email_status = "failed"
                    print(f"EMAIL FAILED for {business.email}: {exc}")

            if args.no_whatsapp:
                whatsapp_status = "disabled"
            elif not business.phone:
                whatsapp_status = "no_recipient"
            elif whatsapp_sender:
                try:
                    whatsapp_sender.send(
                        business.phone,
                        message.whatsapp_text,
                        media_url=os.getenv("TWILIO_MEDIA_URL"),
                    )
                    whatsapp_status = "sent"
                except Exception as exc:
                    whatsapp_status = "failed"
                    print(f"WHATSAPP FAILED for {business.phone}: {exc}")
            else:
                whatsapp_status = "not_configured"

            append_sent(config.SENT_LOG, niche_name, business, message, email_status, whatsapp_status)
            if email_status == "sent" and tracking_pixel_url:
                append_tracking(TRACKING_LOG, tracking_id, niche_name, business, message)
            print(f"SENT {business.name}: email={email_status}, whatsapp={whatsapp_status}")

            if index < len(leads):
                time.sleep(max(args.delay, 0))


def parse_args():
    parser = argparse.ArgumentParser(description="Send configured promotion messages.")
    parser.add_argument("--niche", choices=config.NICHES.keys())
    parser.add_argument("--sheet", type=str)
    parser.add_argument("--limit", type=int)
    parser.add_argument("--delay", type=int, default=config.SEND_DELAY_SECONDS)
    parser.add_argument("--dry-run", action="store_true", help="Preview messages without sending.")
    parser.add_argument("--send", action="store_true", help="Actually send messages. Default is dry-run.")
    parser.add_argument("--check", action="store_true", help="Validate files and credentials.")
    parser.add_argument("--preview", action="store_true", help="Write rendered email HTML to data/previews.")
    parser.add_argument("--email-subject", type=str, help="Override the configured email subject.")
    parser.add_argument("--email-intro", type=str, help="Override the configured email cover letter text.")
    parser.add_argument("--template-url", type=str, help="Override the template/demo URL used in the email.")
    parser.add_argument("--preview-image", type=str, help="Override the image embedded in the email.")
    parser.add_argument("--no-preview-image", action="store_const", const="", dest="preview_image")
    parser.add_argument("--tracking-base-url", type=str, help="Public base URL used for email open tracking pixels.")
    parser.add_argument("--resend", action="store_true")
    parser.add_argument("--no-email", action="store_true")
    parser.add_argument("--no-whatsapp", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
